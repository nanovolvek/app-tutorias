from fastapi import APIRouter, Depends, HTTPException, status, Body, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from typing import List
from app.database import get_db
from app.models.student import Estudiante
from app.models.equipo import Equipo
from app.models.school import Colegio
from app.schemas.estudiante import Estudiante as EstudianteSchema, EstudianteCreate, EstudianteDeleteRequest
from app.models.attendance import AsistenciaEstudiante
from app.models.tickets import TicketEstudiante
from app.models.prueba_diagnostico import PruebaDiagnosticoEstudiante
from app.models.prueba_unidad import PruebaUnidadEstudiante
from app.auth.dependencies import get_current_active_user, get_admin_user, get_tutor_user
from openpyxl import load_workbook
import re
from pydantic import BaseModel

router = APIRouter(prefix="/estudiantes", tags=["estudiantes"])

@router.get("/", response_model=List[EstudianteSchema])
def get_estudiantes(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Obtener estudiantes según el rol del usuario"""
    try:
        if current_user.rol == "admin":
            # Admin puede ver todos los estudiantes
            estudiantes = db.query(Estudiante).options(
                joinedload(Estudiante.equipo).joinedload(Equipo.colegio)
            ).all()
        else:
            # Tutor solo puede ver estudiantes de su equipo
            estudiantes = db.query(Estudiante).options(
                joinedload(Estudiante.equipo).joinedload(Equipo.colegio)
            ).filter(Estudiante.equipo_id == current_user.equipo_id).all()
        
        # Asegurar que activo tenga un valor por defecto si es None
        for estudiante in estudiantes:
            if estudiante.activo is None:
                estudiante.activo = True
        
        return estudiantes
    except Exception as e:
        print(f"Error en get_estudiantes: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al obtener estudiantes: {str(e)}")

@router.get("/{estudiante_id}", response_model=EstudianteSchema)
def get_estudiante(
    estudiante_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Obtener un estudiante específico"""
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Estudiante no encontrado"
        )
    
    # Verificar permisos
    if current_user.rol == "tutor" and estudiante.equipo_id != current_user.equipo_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para ver este estudiante"
        )
    
    return estudiante

@router.post("/", response_model=EstudianteSchema)
def create_estudiante(
    estudiante: EstudianteCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Crear un nuevo estudiante"""
    # Verificar que el equipo existe
    equipo = db.query(Equipo).filter(Equipo.id == estudiante.equipo_id).first()
    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )
    
    # Si es tutor, verificar que solo puede agregar estudiantes a su equipo
    if current_user.rol == "tutor" and estudiante.equipo_id != current_user.equipo_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo puedes agregar estudiantes a tu equipo"
        )
    
    # Verificar que el RUT no esté en uso
    existing_estudiante = db.query(Estudiante).filter(Estudiante.rut == estudiante.rut).first()
    if existing_estudiante:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ya existe un estudiante con este RUT"
        )
    
    try:
        db_estudiante = Estudiante(**estudiante.dict())
        db.add(db_estudiante)
        db.commit()
        db.refresh(db_estudiante)
        return db_estudiante
    except Exception as e:
        db.rollback()
        # Si hay error de ID duplicado, intentar obtener el siguiente ID disponible
        if "llave duplicada" in str(e) or "duplicate key" in str(e):
            # Obtener el máximo ID actual
            max_id = db.query(Estudiante).order_by(Estudiante.id.desc()).first()
            next_id = (max_id.id + 1) if max_id else 1
            
            # Crear el estudiante con ID explícito
            estudiante_data = estudiante.dict()
            db_estudiante = Estudiante(id=next_id, **estudiante_data)
            db.add(db_estudiante)
            db.commit()
            db.refresh(db_estudiante)
            return db_estudiante
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error al crear el estudiante: {str(e)}"
            )

@router.delete("/{estudiante_id}")
def delete_estudiante(
    estudiante_id: int,
    delete_request: EstudianteDeleteRequest = Body(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Eliminar un estudiante o marcarlo como desertor"""
    estudiante = db.query(Estudiante).filter(Estudiante.id == estudiante_id).first()
    if not estudiante:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Estudiante no encontrado"
        )
    
    # Verificar permisos - tutores solo pueden eliminar estudiantes de su equipo
    if current_user.rol == "tutor" and estudiante.equipo_id != current_user.equipo_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo puedes eliminar estudiantes de tu equipo"
        )
    
    if delete_request.es_desercion:
        # Marcar como desertor (eliminación lógica)
        if not delete_request.motivo_desercion:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El motivo de deserción es requerido"
            )
        estudiante.activo = False
        estudiante.motivo_desercion = delete_request.motivo_desercion
        db.commit()
        db.refresh(estudiante)
        return {"message": "Estudiante marcado como desertor exitosamente"}
    else:
        # Eliminación física completa - eliminar todos los registros relacionados
        # Eliminar registros de asistencia
        db.query(AsistenciaEstudiante).filter(AsistenciaEstudiante.estudiante_id == estudiante_id).delete()
        # Eliminar registros de tickets
        db.query(TicketEstudiante).filter(TicketEstudiante.estudiante_id == estudiante_id).delete()
        # Eliminar registros de prueba diagnóstico
        db.query(PruebaDiagnosticoEstudiante).filter(PruebaDiagnosticoEstudiante.estudiante_id == estudiante_id).delete()
        # Eliminar registros de prueba unidad
        db.query(PruebaUnidadEstudiante).filter(PruebaUnidadEstudiante.estudiante_id == estudiante_id).delete()
        # Eliminar el estudiante
        db.delete(estudiante)
        db.commit()
        return {"message": "Estudiante eliminado completamente de la base de datos"}

def validate_rut(rut: str) -> bool:
    """Validar formato de RUT chileno: XX.XXX.XXX-X o X.XXX.XXX-X"""
    pattern = r'^(\d{1,2}\.\d{3}\.\d{3}-[\dkK])$'
    return bool(re.match(pattern, rut.strip()))

@router.post("/import")
async def import_estudiantes(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Importar estudiantes desde un archivo Excel"""
    
    # Verificar que el archivo sea Excel
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser un Excel (.xlsx o .xls)"
        )
    
    try:
        # Leer el archivo Excel
        contents = await file.read()
        from io import BytesIO
        workbook = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        
        # Validar encabezados esperados
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['RUT', 'Nombre', 'Apellido', 'Curso', 'Equipo ID', 'Nombre Apoderado', 'Contacto Apoderado', 'Observaciones']
        
        # Verificar que los encabezados estén presentes (case insensitive)
        headers_lower = [str(h).strip().lower() if h else '' for h in headers]
        expected_lower = [h.lower() for h in expected_headers]
        
        missing_headers = []
        for expected in expected_lower:
            if expected not in headers_lower:
                missing_headers.append(expected)
        
        if missing_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Faltan las siguientes columnas requeridas: {', '.join(missing_headers)}"
            )
        
        # Obtener índices de columnas
        header_map = {}
        for idx, header in enumerate(headers, start=1):
            if header:
                header_lower = str(header).strip().lower()
                if header_lower in expected_lower:
                    header_map[expected_lower.index(header_lower)] = idx
        
        # Procesar filas
        created = 0
        errors = []
        existing_ruts = set()
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Saltar filas vacías
            if not any(row):
                continue
            
            try:
                # Extraer datos según índices
                rut = str(row[header_map[0] - 1]).strip() if row[header_map[0] - 1] else None
                nombre = str(row[header_map[1] - 1]).strip() if row[header_map[1] - 1] else None
                apellido = str(row[header_map[2] - 1]).strip() if row[header_map[2] - 1] else None
                curso = str(row[header_map[3] - 1]).strip() if row[header_map[3] - 1] else None
                equipo_id = row[header_map[4] - 1]
                nombre_apoderado = str(row[header_map[5] - 1]).strip() if len(header_map) > 5 and row[header_map[5] - 1] else None
                contacto_apoderado = str(row[header_map[6] - 1]).strip() if len(header_map) > 6 and row[header_map[6] - 1] else None
                observaciones = str(row[header_map[7] - 1]).strip() if len(header_map) > 7 and row[header_map[7] - 1] else None
                
                # Validaciones
                if not rut:
                    errors.append(f"Fila {row_num}: RUT es requerido")
                    continue
                
                if not validate_rut(rut):
                    errors.append(f"Fila {row_num}: RUT '{rut}' no tiene el formato correcto (debe ser XX.XXX.XXX-X o X.XXX.XXX-X)")
                    continue
                
                if not nombre:
                    errors.append(f"Fila {row_num}: Nombre es requerido")
                    continue
                
                if not apellido:
                    errors.append(f"Fila {row_num}: Apellido es requerido")
                    continue
                
                if not curso:
                    errors.append(f"Fila {row_num}: Curso es requerido")
                    continue
                
                if not equipo_id:
                    errors.append(f"Fila {row_num}: Equipo ID es requerido")
                    continue
                
                try:
                    equipo_id = int(equipo_id)
                except (ValueError, TypeError):
                    errors.append(f"Fila {row_num}: Equipo ID '{equipo_id}' debe ser un número")
                    continue
                
                # Verificar que el equipo existe
                equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()
                if not equipo:
                    errors.append(f"Fila {row_num}: Equipo ID {equipo_id} no existe")
                    continue
                
                # Verificar permisos - tutores solo pueden agregar a su equipo
                if current_user.rol == "tutor" and equipo_id != current_user.equipo_id:
                    errors.append(f"Fila {row_num}: Solo puedes agregar estudiantes a tu equipo (ID: {current_user.equipo_id})")
                    continue
                
                # Verificar RUT duplicado en el archivo
                if rut in existing_ruts:
                    errors.append(f"Fila {row_num}: RUT '{rut}' está duplicado en el archivo")
                    continue
                existing_ruts.add(rut)
                
                # Verificar que el RUT no esté en uso en la base de datos
                existing_estudiante = db.query(Estudiante).filter(Estudiante.rut == rut).first()
                if existing_estudiante:
                    errors.append(f"Fila {row_num}: Ya existe un estudiante con RUT '{rut}'")
                    continue
                
                # Crear estudiante
                estudiante_data = {
                    "rut": rut,
                    "nombre": nombre,
                    "apellido": apellido,
                    "curso": curso,
                    "equipo_id": equipo_id,
                    "nombre_apoderado": nombre_apoderado if nombre_apoderado else None,
                    "contacto_apoderado": contacto_apoderado if contacto_apoderado else None,
                    "observaciones": observaciones if observaciones else None
                }
                
                try:
                    db_estudiante = Estudiante(**estudiante_data)
                    db.add(db_estudiante)
                    db.commit()
                    db.refresh(db_estudiante)
                    created += 1
                except Exception as e:
                    db.rollback()
                    # Si hay error de ID duplicado, intentar obtener el siguiente ID disponible
                    if "llave duplicada" in str(e) or "duplicate key" in str(e) or "UniqueViolation" in str(e):
                        try:
                            # Obtener el máximo ID actual
                            max_id = db.query(Estudiante).order_by(Estudiante.id.desc()).first()
                            next_id = (max_id.id + 1) if max_id else 1
                            
                            # Crear el estudiante con ID explícito
                            db_estudiante = Estudiante(id=next_id, **estudiante_data)
                            db.add(db_estudiante)
                            db.commit()
                            db.refresh(db_estudiante)
                            created += 1
                        except Exception as e2:
                            db.rollback()
                            errors.append(f"Fila {row_num}: Error al crear estudiante (ID duplicado y fallo al corregir) - {str(e2)}")
                    else:
                        errors.append(f"Fila {row_num}: Error al procesar - {str(e)}")
                
            except Exception as e:
                db.rollback()
                errors.append(f"Fila {row_num}: Error al procesar - {str(e)}")
                continue
        
        return {
            "message": f"Importación completada: {created} estudiantes creados",
            "created": created,
            "errors": errors,
            "total_errors": len(errors)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo Excel: {str(e)}"
        )
