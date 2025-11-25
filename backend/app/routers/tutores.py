from fastapi import APIRouter, Depends, HTTPException, status, Body, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from typing import List
from app.database import get_db
from app.models.tutor import Tutor
from app.models.equipo import Equipo
from app.models.school import Colegio
from app.schemas.tutor import Tutor as TutorSchema, TutorCreate, TutorDeleteRequest
from app.models.attendance import AsistenciaTutor
from app.auth.dependencies import get_current_active_user, get_admin_user, get_tutor_user
from openpyxl import load_workbook
from io import BytesIO
import re

router = APIRouter(prefix="/tutores", tags=["tutores"])

@router.get("/", response_model=List[TutorSchema])
def get_tutores(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Obtener tutores según el rol del usuario"""
    try:
        if current_user.rol == "admin":
            # Admin puede ver todos los tutores
            tutores = db.query(Tutor).options(
                joinedload(Tutor.equipo).joinedload(Equipo.colegio)
            ).all()
        else:
            # Tutor solo puede ver tutores de su equipo
            tutores = db.query(Tutor).options(
                joinedload(Tutor.equipo).joinedload(Equipo.colegio)
            ).filter(Tutor.equipo_id == current_user.equipo_id).all()
        
        # Asegurar que activo tenga un valor por defecto si es None
        for tutor in tutores:
            if tutor.activo is None:
                tutor.activo = True
        
        return tutores
    except Exception as e:
        print(f"Error en get_tutores: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al obtener tutores: {str(e)}")

@router.get("/{tutor_id}", response_model=TutorSchema)
def get_tutor(
    tutor_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Obtener un tutor específico"""
    tutor = db.query(Tutor).filter(Tutor.id == tutor_id).first()
    if not tutor:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tutor no encontrado"
        )
    
    # Verificar permisos
    if current_user.rol == "tutor" and tutor.equipo_id != current_user.equipo_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para ver este tutor"
        )
    
    return tutor

@router.post("/", response_model=TutorSchema)
def create_tutor(
    tutor: TutorCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
):
    """Crear un nuevo tutor (solo administradores)"""
    # Verificar que el equipo existe
    equipo = db.query(Equipo).filter(Equipo.id == tutor.equipo_id).first()
    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )
    
    # Verificar que el email no esté en uso
    existing_tutor = db.query(Tutor).filter(Tutor.email == tutor.email).first()
    if existing_tutor:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ya existe un tutor con este email"
        )
    
    try:
        db_tutor = Tutor(**tutor.dict())
        db.add(db_tutor)
        db.commit()
        db.refresh(db_tutor)
        return db_tutor
    except Exception as e:
        db.rollback()
        # Si hay error de ID duplicado, intentar obtener el siguiente ID disponible
        if "llave duplicada" in str(e) or "duplicate key" in str(e):
            # Obtener el máximo ID actual
            max_id = db.query(Tutor).order_by(Tutor.id.desc()).first()
            next_id = (max_id.id + 1) if max_id else 1
            
            # Crear el tutor con ID explícito
            tutor_data = tutor.dict()
            db_tutor = Tutor(id=next_id, **tutor_data)
            db.add(db_tutor)
            db.commit()
            db.refresh(db_tutor)
            return db_tutor
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error al crear el tutor: {str(e)}"
            )

@router.delete("/{tutor_id}")
def delete_tutor(
    tutor_id: int,
    delete_request: TutorDeleteRequest = Body(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
):
    """Eliminar un tutor o marcarlo como desertor (solo administradores)"""
    tutor = db.query(Tutor).filter(Tutor.id == tutor_id).first()
    if not tutor:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tutor no encontrado"
        )
    
    if delete_request.es_desercion:
        # Marcar como desertor (eliminación lógica)
        if not delete_request.motivo_desercion:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El motivo de deserción es requerido"
            )
        tutor.activo = False
        tutor.motivo_desercion = delete_request.motivo_desercion
        db.commit()
        db.refresh(tutor)
        return {"message": "Tutor marcado como desertor exitosamente"}
    else:
        # Eliminación física completa - eliminar todos los registros relacionados
        # Eliminar registros de asistencia
        db.query(AsistenciaTutor).filter(AsistenciaTutor.tutor_id == tutor_id).delete()
        # Eliminar el tutor
        db.delete(tutor)
        db.commit()
        return {"message": "Tutor eliminado completamente de la base de datos"}

@router.post("/import")
async def import_tutores(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(get_admin_user)
):
    """Importar tutores desde un archivo Excel (solo administradores)"""
    
    # Verificar que el archivo sea Excel
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser un Excel (.xlsx o .xls)"
        )
    
    try:
        # Leer el archivo Excel
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        
        # Validar encabezados esperados
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['Nombre', 'Apellido', 'Email', 'Equipo ID']
        
        # Verificar que los encabezados estén presentes (case insensitive)
        headers_lower = [str(h).strip().lower() if h else '' for h in headers]
        expected_lower = [h.lower() for h in expected_headers]
        
        missing_headers = []
        for expected in expected_lower:
            if expected not in headers_lower:
                missing_headers.append(expected)
        
        if missing_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Faltan las siguientes columnas requeridas: {', '.join(missing_headers)}"
            )
        
        # Obtener índices de columnas
        header_map = {}
        for idx, header in enumerate(headers, start=1):
            if header:
                header_lower = str(header).strip().lower()
                if header_lower in expected_lower:
                    header_map[expected_lower.index(header_lower)] = idx
        
        # Procesar filas
        created = 0
        errors = []
        existing_emails = set()
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Saltar filas vacías
            if not any(row):
                continue
            
            try:
                # Extraer datos según índices
                nombre = str(row[header_map[0] - 1]).strip() if row[header_map[0] - 1] else None
                apellido = str(row[header_map[1] - 1]).strip() if row[header_map[1] - 1] else None
                email = str(row[header_map[2] - 1]).strip() if row[header_map[2] - 1] else None
                equipo_id = row[header_map[3] - 1]
                
                # Validaciones
                if not nombre:
                    errors.append(f"Fila {row_num}: Nombre es requerido")
                    continue
                
                if not apellido:
                    errors.append(f"Fila {row_num}: Apellido es requerido")
                    continue
                
                if not email:
                    errors.append(f"Fila {row_num}: Email es requerido")
                    continue
                
                # Validar formato de email básico
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, email):
                    errors.append(f"Fila {row_num}: Email '{email}' no tiene un formato válido")
                    continue
                
                if not equipo_id:
                    errors.append(f"Fila {row_num}: Equipo ID es requerido")
                    continue
                
                try:
                    equipo_id = int(equipo_id)
                except (ValueError, TypeError):
                    errors.append(f"Fila {row_num}: Equipo ID '{equipo_id}' debe ser un número")
                    continue
                
                # Verificar que el equipo existe
                equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()
                if not equipo:
                    errors.append(f"Fila {row_num}: Equipo ID {equipo_id} no existe")
                    continue
                
                # Verificar email duplicado en el archivo
                if email in existing_emails:
                    errors.append(f"Fila {row_num}: Email '{email}' está duplicado en el archivo")
                    continue
                existing_emails.add(email)
                
                # Verificar que el email no esté en uso en la base de datos
                existing_tutor = db.query(Tutor).filter(Tutor.email == email).first()
                if existing_tutor:
                    errors.append(f"Fila {row_num}: Ya existe un tutor con email '{email}'")
                    continue
                
                # Crear tutor
                tutor_data = {
                    "nombre": nombre,
                    "apellido": apellido,
                    "email": email,
                    "equipo_id": equipo_id
                }
                
                db_tutor = Tutor(**tutor_data)
                db.add(db_tutor)
                db.commit()
                db.refresh(db_tutor)
                created += 1
                
            except Exception as e:
                db.rollback()
                errors.append(f"Fila {row_num}: Error al procesar - {str(e)}")
                continue
        
        return {
            "message": f"Importación completada: {created} tutores creados",
            "created": created,
            "errors": errors,
            "total_errors": len(errors)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar el archivo Excel: {str(e)}"
        )
